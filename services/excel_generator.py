"""Generate institutional-grade Excel workbook with up to 11 tabs and charts."""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR

# --- Style Constants ---
NAVY = "1B2A4A"
DARK_BLUE = "2C3E6B"
GOLD = "C8A951"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
MED_GRAY = "D9D9D9"

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color=WHITE)
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, size=11, color=NAVY)
SUBHEADER_FILL = PatternFill(start_color=MED_GRAY, end_color=MED_GRAY, fill_type="solid")
NORMAL_FONT = Font(name="Calibri", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color=NAVY)
ALT_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
CURRENCY_FMT = '$#,##0'
PCT_FMT = '0.00%'
THIN_BORDER = Border(
    left=Side(style="thin", color=MED_GRAY),
    right=Side(style="thin", color=MED_GRAY),
    top=Side(style="thin", color=MED_GRAY),
    bottom=Side(style="thin", color=MED_GRAY),
)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


def _hdr_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _data_row(ws, row, max_col, alt=False):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        if alt:
            cell.fill = ALT_FILL


def _title(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = TITLE_FONT


def _widths(ws, max_col, w=15):
    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def generate_excel(pro_forma: dict, market_data: dict, job_id: str,
                   ml_valuation: dict = None,
                   lease_analysis: dict = None,
                   rent_prediction: dict = None,
                   sensitivity: dict = None) -> str:
    wb = Workbook()
    inp = pro_forma["inputs"]
    deal = inp["deal"]
    derived = inp["derived"]
    metrics = pro_forma["metrics"]
    pf = pro_forma["pro_forma"]
    su = pro_forma["sources_uses"]
    amort = pro_forma["amortization_annual"]
    rev = pro_forma["reversion"]

    # ========== TAB 1: Summary ==========
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = NAVY

    _title(ws, 1, 1, "Investment Summary")
    r = 3
    snap = [
        ("Property", derived["property_name"]),
        ("Type", deal["property_type"]),
        ("Address", deal["address"]),
        ("Year Built", deal["year_built"]),
        ("Units / SF", f"{deal['total_units']} units / {deal['total_sf']:,.0f} SF"),
        ("Purchase Price", deal["purchase_price"]),
        ("Price / Unit", derived["price_per_unit"]),
        ("Current NOI", deal["current_noi"]),
        ("In-Place Rent", f"${deal['in_place_rent']:,.0f}/mo"),
        ("Market Rent", f"${deal['market_rent']:,.0f}/mo"),
        ("Occupancy", deal["occupancy"]),
        ("Total CapEx", derived["total_capex"]),
    ]
    for label, val in snap:
        ws.cell(row=r, column=1, value=label).font = SUBHEADER_FONT
        c = ws.cell(row=r, column=2, value=val)
        c.font = NORMAL_FONT
        if isinstance(val, float) and val < 1 and val > 0:
            c.number_format = PCT_FMT
        elif isinstance(val, (int, float)) and val > 1000:
            c.number_format = CURRENCY_FMT
        r += 1

    r += 1
    _title(ws, r, 1, "Return Metrics")
    r += 1
    for c_idx, h in enumerate(["Metric", "Value"], 1):
        ws.cell(row=r, column=c_idx, value=h)
    _hdr_row(ws, r, 2)
    r += 1

    ret_items = [
        ("Levered IRR", metrics["levered_irr"], True),
        ("After-Tax IRR", metrics.get("after_tax_irr"), True),
        ("Unlevered IRR", metrics["unlevered_irr"], True),
        ("Equity Multiple", metrics["equity_multiple"], False),
        ("After-Tax Equity Multiple", metrics.get("after_tax_equity_multiple"), False),
        ("Cash-on-Cash (Yr 1)", metrics["cash_on_cash_yr1"], True),
        ("After-Tax CoC (Yr 1)", metrics.get("cash_on_cash_yr1_after_tax"), True),
        ("DSCR (Yr 1)", metrics["dscr_yr1"], False),
        ("Stabilized DSCR", metrics["stabilized_dscr"], False),
        ("Yield on Cost", metrics["yield_on_cost"], True),
        ("Stabilized YOC", metrics["stabilized_yoc"], True),
        ("Going-In Cap Rate", metrics["going_in_cap_rate"], True),
        ("Exit Cap Rate", metrics["exit_cap_rate"], True),
    ]
    for label, val, is_pct in ret_items:
        ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
        c = ws.cell(row=r, column=2)
        if val is not None and is_pct:
            c.value = val
            c.number_format = PCT_FMT
        elif val is not None:
            c.value = round(val, 2)
            c.number_format = '0.00"x"'
        else:
            c.value = "N/A"
        c.font = NORMAL_FONT
        r += 1

    # NOI chart
    if pf:
        r += 2
        ws.cell(row=r, column=1, value="Year")
        ws.cell(row=r, column=2, value="NOI")
        for i, yr in enumerate(pf):
            ws.cell(row=r+1+i, column=1, value=f"Year {yr['year']}")
            ws.cell(row=r+1+i, column=2, value=yr["noi"])
        chart = BarChart()
        chart.title = "Net Operating Income"
        chart.style = 10
        chart.width = 20
        chart.height = 12
        chart.add_data(Reference(ws, min_col=2, min_row=r, max_row=r+len(pf)), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=r+1, max_row=r+len(pf)))
        chart.series[0].graphicalProperties.solidFill = NAVY
        ws.add_chart(chart, f"D3")
    _widths(ws, 5)

    # ========== TAB 2: Sources & Uses ==========
    ws2 = wb.create_sheet("Sources & Uses")
    ws2.sheet_properties.tabColor = DARK_BLUE
    _title(ws2, 1, 1, "Sources & Uses of Funds")

    r = 3
    for h_idx, h in enumerate(["SOURCES", "Amount", "% of Total"], 1):
        ws2.cell(row=r, column=h_idx, value=h)
    _hdr_row(ws2, r, 3)
    r += 1
    total_s = su["sources"]["total"]
    for label, key in [("Senior Debt", "senior_debt"), ("Sponsor Equity", "sponsor_equity")]:
        ws2.cell(row=r, column=1, value=label).font = NORMAL_FONT
        ws2.cell(row=r, column=2, value=su["sources"][key]).number_format = CURRENCY_FMT
        ws2.cell(row=r, column=3, value=su["sources"][key]/total_s if total_s else 0).number_format = PCT_FMT
        r += 1
    ws2.cell(row=r, column=1, value="Total Sources").font = SUBHEADER_FONT
    ws2.cell(row=r, column=2, value=total_s).number_format = CURRENCY_FMT
    ws2.cell(row=r, column=2).font = SUBHEADER_FONT

    r += 2
    for h_idx, h in enumerate(["USES", "Amount", "% of Total"], 1):
        ws2.cell(row=r, column=h_idx, value=h)
    _hdr_row(ws2, r, 3)
    r += 1
    total_u = su["uses"]["total"]
    uses_items = [
        ("Purchase Price", su["uses"]["purchase_price"]),
        ("Closing Costs", su["uses"]["closing_costs"]),
    ]
    if su["uses"]["deferred_maintenance"] > 0:
        uses_items.append(("Deferred Maintenance", su["uses"]["deferred_maintenance"]))
    if su["uses"]["planned_capex"] > 0:
        desc = deal.get("capex_description", "")
        label = f"Planned CapEx ({desc})" if desc else "Planned CapEx"
        uses_items.append((label, su["uses"]["planned_capex"]))

    for label, val in uses_items:
        ws2.cell(row=r, column=1, value=label).font = NORMAL_FONT
        ws2.cell(row=r, column=2, value=val).number_format = CURRENCY_FMT
        ws2.cell(row=r, column=3, value=val/total_u if total_u else 0).number_format = PCT_FMT
        r += 1
    ws2.cell(row=r, column=1, value="Total Uses").font = SUBHEADER_FONT
    ws2.cell(row=r, column=2, value=total_u).number_format = CURRENCY_FMT
    ws2.cell(row=r, column=2).font = SUBHEADER_FONT
    _widths(ws2, 3)

    # ========== TAB 3: Pro Forma ==========
    ws3 = wb.create_sheet("Pro Forma")
    ws3.sheet_properties.tabColor = DARK_BLUE
    _title(ws3, 1, 1, "10-Year Pro Forma")

    headers = [""] + [f"Year {i}" for i in range(1, 11)]
    r = 3
    for c, h in enumerate(headers, 1):
        ws3.cell(row=r, column=c, value=h)
    _hdr_row(ws3, r, len(headers))

    rows = [
        ("Rent / Unit ($/mo)", "rent_per_unit", CURRENCY_FMT),
        ("Revenue Growth Used", "revenue_growth_used", '0.00"%"'),
        ("Occupancy", "occupancy", PCT_FMT),
        ("Gross Potential Rent", "gross_potential_rent", CURRENCY_FMT),
        ("Less: Vacancy", "vacancy_loss", CURRENCY_FMT),
        ("Other Income", "other_income", CURRENCY_FMT),
        ("Effective Gross Income", "effective_gross_income", CURRENCY_FMT),
        ("", None, None),
        ("Management Fee", "management_fee", CURRENCY_FMT),
        ("Property Tax", "property_tax", CURRENCY_FMT),
        ("Insurance", "insurance", CURRENCY_FMT),
        ("Utilities", "utilities", CURRENCY_FMT),
        ("Repairs & Maintenance", "repairs_maintenance", CURRENCY_FMT),
        ("General & Admin", "general_admin", CURRENCY_FMT),
        ("Other Expenses", "other_expenses", CURRENCY_FMT),
        ("Replacement Reserves", "replacement_reserves", CURRENCY_FMT),
        ("Total Expenses", "total_expenses", CURRENCY_FMT),
        ("", None, None),
        ("Net Operating Income", "noi", CURRENCY_FMT),
        ("Debt Service", "debt_service", CURRENCY_FMT),
        ("Before-Tax Cash Flow", "btcf", CURRENCY_FMT),
        ("", None, None),
        ("Interest Expense", "interest_expense", CURRENCY_FMT),
        ("Depreciation", "depreciation", CURRENCY_FMT),
        ("Taxable Income", "taxable_income", CURRENCY_FMT),
        ("Tax Liability", "tax_liability", CURRENCY_FMT),
        ("After-Tax Cash Flow", "atcf", CURRENCY_FMT),
    ]
    bold_rows = {"Effective Gross Income", "Total Expenses", "Net Operating Income",
                 "Before-Tax Cash Flow", "After-Tax Cash Flow"}

    for i, (label, key, fmt) in enumerate(rows):
        r += 1
        ws3.cell(row=r, column=1, value=label)
        if label in bold_rows:
            ws3.cell(row=r, column=1).font = SUBHEADER_FONT
        else:
            ws3.cell(row=r, column=1).font = NORMAL_FONT
        if key:
            for yr_idx, yr_data in enumerate(pf):
                c = ws3.cell(row=r, column=yr_idx + 2, value=yr_data[key])
                if fmt:
                    c.number_format = fmt
                c.font = NORMAL_FONT
        _data_row(ws3, r, len(headers), alt=(i % 2 == 0))

    _widths(ws3, len(headers), 14)
    ws3.column_dimensions["A"].width = 24

    # ========== TAB 4: Loan ==========
    ws4 = wb.create_sheet("Loan")
    ws4.sheet_properties.tabColor = DARK_BLUE
    _title(ws4, 1, 1, "Loan Amortization")

    r = 3
    for label, val, fmt in [
        ("Loan Amount", derived["loan_amount"], CURRENCY_FMT),
        ("LTV", deal["ltv"], PCT_FMT),
        ("Interest Rate", deal["interest_rate"], PCT_FMT),
        ("Amortization", f"{deal['amortization_years']} years", None),
        ("Loan Term", f"{deal['loan_term_years']} years", None),
        ("IO Period", f"{deal['io_period_years']} years", None),
    ]:
        ws4.cell(row=r, column=1, value=label).font = SUBHEADER_FONT
        c = ws4.cell(row=r, column=2, value=val)
        if fmt:
            c.number_format = fmt
        c.font = NORMAL_FONT
        r += 1

    if amort:
        r += 1
        hdrs = ["Year", "Total Payment", "Principal", "Interest", "Ending Balance"]
        for c_idx, h in enumerate(hdrs, 1):
            ws4.cell(row=r, column=c_idx, value=h)
        _hdr_row(ws4, r, len(hdrs))
        r += 1
        for i, yr in enumerate(amort):
            ws4.cell(row=r, column=1, value=yr["year"]).font = NORMAL_FONT
            for c_idx, key in enumerate(["total_payment", "total_principal", "total_interest", "ending_balance"], 2):
                ws4.cell(row=r, column=c_idx, value=yr[key]).number_format = CURRENCY_FMT
            _data_row(ws4, r, len(hdrs), alt=(i % 2 == 0))
            r += 1

        chart = LineChart()
        chart.title = "Loan Balance"
        chart.style = 10
        chart.width = 18
        chart.height = 10
        ds = r - len(amort)
        chart.add_data(Reference(ws4, min_col=5, min_row=ds-1, max_row=r-1), titles_from_data=True)
        chart.set_categories(Reference(ws4, min_col=1, min_row=ds, max_row=r-1))
        ws4.add_chart(chart, f"A{r+1}")
    _widths(ws4, 5)

    # ========== TAB 5: Comps ==========
    ws5 = wb.create_sheet("Comps")
    ws5.sheet_properties.tabColor = DARK_BLUE
    _title(ws5, 1, 1, "Comparable Sales")

    comps = market_data.get("comps", {}).get("comps", [])
    r = 3
    comp_h = ["#", "Property", "Price/Unit", "Cap Rate", "Year", "Units"]
    for c_idx, h in enumerate(comp_h, 1):
        ws5.cell(row=r, column=c_idx, value=h)
    _hdr_row(ws5, r, len(comp_h))
    r += 1

    for i, comp in enumerate(comps[:7], 1):
        if not isinstance(comp, dict):
            continue
        ws5.cell(row=r, column=1, value=i).font = NORMAL_FONT
        ws5.cell(row=r, column=2, value=comp.get("name", comp.get("source", f"Comp {i}"))).font = NORMAL_FONT
        ppu = comp.get("price_per_unit", "")
        if isinstance(ppu, (int, float)):
            ws5.cell(row=r, column=3, value=ppu).number_format = CURRENCY_FMT
        else:
            ws5.cell(row=r, column=3, value=ppu)
        cap = comp.get("cap_rate", "")
        if isinstance(cap, (int, float)):
            ws5.cell(row=r, column=4, value=cap/100 if cap > 1 else cap).number_format = PCT_FMT
        else:
            ws5.cell(row=r, column=4, value=cap)
        ws5.cell(row=r, column=5, value=comp.get("year", "")).font = NORMAL_FONT
        ws5.cell(row=r, column=6, value=comp.get("units", "")).font = NORMAL_FONT
        _data_row(ws5, r, len(comp_h), alt=(i % 2 == 0))
        r += 1

    # Stats
    ppus = [c["price_per_unit"] for c in comps if isinstance(c, dict) and isinstance(c.get("price_per_unit"), (int, float))]
    if ppus:
        r += 1
        ws5.cell(row=r, column=2, value="Average").font = SUBHEADER_FONT
        ws5.cell(row=r, column=3, value=round(sum(ppus)/len(ppus))).number_format = CURRENCY_FMT
    _widths(ws5, len(comp_h))

    # ========== TAB 6: Sensitivity ==========
    ws6 = wb.create_sheet("Sensitivity")
    ws6.sheet_properties.tabColor = DARK_BLUE
    _title(ws6, 1, 1, "Sensitivity Analysis")

    # Use pre-computed sensitivity data if available, else compute exit cap only
    if sensitivity and "exit_cap" in sensitivity:
        exit_sens = sensitivity["exit_cap"]
    else:
        from models.metrics import sensitivity_table_exit_cap
        exit_sens = sensitivity_table_exit_cap(
            base_noi_at_exit=rev["forward_noi"],
            base_exit_cap=rev["exit_cap_rate"],
            equity_invested=derived["equity_required"],
            annual_cfs=pro_forma["annual_btcfs"][:rev["exit_year"]],
            sale_costs_pct=deal["sale_costs_pct"],
            loan_balance_at_exit=rev["loan_balance"],
        )

    r = 3
    ws6.cell(row=r, column=1, value="Exit Cap Rate Sensitivity").font = SUBHEADER_FONT
    r += 1
    sh = ["Exit Cap", "Sale Price", "IRR", "Equity Multiple"]
    for c_idx, h in enumerate(sh, 1):
        ws6.cell(row=r, column=c_idx, value=h)
    _hdr_row(ws6, r, len(sh))
    r += 1

    for i, row_data in enumerate(exit_sens):
        ws6.cell(row=r, column=1, value=row_data["exit_cap_rate"]/100).number_format = PCT_FMT
        ws6.cell(row=r, column=2, value=row_data["sale_price"]).number_format = CURRENCY_FMT
        irr_c = ws6.cell(row=r, column=3)
        if row_data["irr"] is not None:
            irr_c.value = row_data["irr"] / 100
            irr_c.number_format = PCT_FMT
            if row_data["irr"] >= 15:
                irr_c.fill = GREEN_FILL
            elif row_data["irr"] >= 10:
                irr_c.fill = YELLOW_FILL
            else:
                irr_c.fill = RED_FILL
        else:
            irr_c.value = "N/A"
        ws6.cell(row=r, column=4, value=row_data["equity_multiple"])
        _data_row(ws6, r, len(sh), alt=(i % 2 == 0))
        r += 1

    # Interest Rate Sensitivity (Fix #7)
    if sensitivity and "interest_rate" in sensitivity:
        r += 2
        ws6.cell(row=r, column=1, value="Interest Rate Sensitivity").font = SUBHEADER_FONT
        r += 1
        sh2 = ["Interest Rate", "IRR", "Equity Multiple", "DSCR"]
        for c_idx, h in enumerate(sh2, 1):
            ws6.cell(row=r, column=c_idx, value=h)
        _hdr_row(ws6, r, len(sh2))
        r += 1
        for i, row_data in enumerate(sensitivity["interest_rate"]):
            ws6.cell(row=r, column=1, value=row_data["interest_rate"]/100).number_format = PCT_FMT
            irr_c = ws6.cell(row=r, column=2)
            if row_data["irr"] is not None:
                irr_c.value = row_data["irr"] / 100
                irr_c.number_format = PCT_FMT
            else:
                irr_c.value = "N/A"
            ws6.cell(row=r, column=3, value=row_data.get("equity_multiple"))
            dscr_c = ws6.cell(row=r, column=4, value=row_data.get("dscr"))
            if row_data.get("dscr") and row_data["dscr"] < 1.0:
                dscr_c.fill = RED_FILL
            elif row_data.get("dscr") and row_data["dscr"] < 1.25:
                dscr_c.fill = YELLOW_FILL
            _data_row(ws6, r, len(sh2), alt=(i % 2 == 0))
            r += 1

    # Rent Growth Sensitivity (Fix #7)
    if sensitivity and "rent_growth" in sensitivity:
        r += 2
        ws6.cell(row=r, column=1, value="Rent Growth Sensitivity").font = SUBHEADER_FONT
        r += 1
        sh3 = ["Rent Growth", "IRR", "Equity Multiple", "Stab. YOC"]
        for c_idx, h in enumerate(sh3, 1):
            ws6.cell(row=r, column=c_idx, value=h)
        _hdr_row(ws6, r, len(sh3))
        r += 1
        for i, row_data in enumerate(sensitivity["rent_growth"]):
            ws6.cell(row=r, column=1, value=row_data["rent_growth"]/100).number_format = PCT_FMT
            irr_c = ws6.cell(row=r, column=2)
            if row_data["irr"] is not None:
                irr_c.value = row_data["irr"] / 100
                irr_c.number_format = PCT_FMT
            else:
                irr_c.value = "N/A"
            ws6.cell(row=r, column=3, value=row_data.get("equity_multiple"))
            yoc_c = ws6.cell(row=r, column=4)
            if row_data.get("stabilized_yoc") is not None:
                yoc_c.value = row_data["stabilized_yoc"] / 100
                yoc_c.number_format = PCT_FMT
            _data_row(ws6, r, len(sh3), alt=(i % 2 == 0))
            r += 1

    # Purchase Price Sensitivity
    if sensitivity and "purchase_price" in sensitivity:
        r += 2
        ws6.cell(row=r, column=1, value="Purchase Price Sensitivity").font = SUBHEADER_FONT
        r += 1
        sh4 = ["Price Change", "Purchase Price", "IRR", "Equity Multiple"]
        for c_idx, h in enumerate(sh4, 1):
            ws6.cell(row=r, column=c_idx, value=h)
        _hdr_row(ws6, r, len(sh4))
        r += 1
        for i, row_data in enumerate(sensitivity["purchase_price"]):
            ws6.cell(row=r, column=1, value=row_data["price_change"]).font = NORMAL_FONT
            ws6.cell(row=r, column=2, value=row_data["purchase_price"]).number_format = CURRENCY_FMT
            irr_c = ws6.cell(row=r, column=3)
            if row_data["irr"] is not None:
                irr_c.value = row_data["irr"] / 100
                irr_c.number_format = PCT_FMT
            else:
                irr_c.value = "N/A"
            ws6.cell(row=r, column=4, value=row_data.get("equity_multiple"))
            _data_row(ws6, r, len(sh4), alt=(i % 2 == 0))
            r += 1

    _widths(ws6, 4)

    # ========== TAB 7: Market ==========
    ws7 = wb.create_sheet("Market")
    ws7.sheet_properties.tabColor = DARK_BLUE
    _title(ws7, 1, 1, "Market Overview")

    r = 3
    demo = market_data.get("demographics", {})
    ws7.cell(row=r, column=1, value="Location").font = SUBHEADER_FONT
    ws7.cell(row=r, column=2, value=f"{derived['city']}, {derived['state']}").font = NORMAL_FONT
    r += 2

    summary = demo.get("summary", "Market data not available.")
    c = ws7.cell(row=r, column=1, value=summary)
    c.font = NORMAL_FONT
    c.alignment = Alignment(wrap_text=True)
    ws7.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 2

    # Structured market data from APIs
    ws7.cell(row=r, column=1, value="Market Indicators").font = SUBHEADER_FONT
    ws7.cell(row=r, column=1).fill = HEADER_FILL
    ws7.cell(row=r, column=1).font = HEADER_FONT
    ws7.cell(row=r, column=2).fill = HEADER_FILL
    ws7.cell(row=r, column=3).fill = HEADER_FILL
    r += 1

    cap_data = market_data.get("cap_rates", {})
    rent_data = market_data.get("rent_trends", {})
    structured = demo.get("structured", {})

    indicators = [
        ("Market Cap Rate", cap_data.get("average_cap_rate"), "%"),
        ("10-Year Treasury", cap_data.get("treasury_10yr"), "%"),
        ("30-Year Mortgage", cap_data.get("mortgage_30yr") if "mortgage_30yr" in cap_data else None, "%"),
        ("Avg Rent Growth (CPI Shelter)", rent_data.get("average_growth"), "%"),
        ("State Population", structured.get("population"), "#"),
        ("Median Household Income", structured.get("median_income"), "$"),
        ("Median Gross Rent", structured.get("median_rent"), "$"),
        ("Unemployment Rate", structured.get("unemployment_rate"), "%"),
        ("Housing Vacancy Rate", structured.get("vacancy_rate"), "%"),
        ("Renter-Occupied %", structured.get("renter_pct"), "%"),
    ]

    for label, val, fmt_type in indicators:
        ws7.cell(row=r, column=1, value=label).font = NORMAL_FONT
        vc = ws7.cell(row=r, column=2)
        if val is not None:
            if fmt_type == "%":
                vc.value = val / 100
                vc.number_format = PCT_FMT
            elif fmt_type == "$":
                vc.value = val
                vc.number_format = CURRENCY_FMT
            else:
                vc.value = val
                vc.number_format = '#,##0'
        else:
            vc.value = "N/A"
        vc.font = NORMAL_FONT
        _data_row(ws7, r, 2)
        r += 1

    # Data sources
    r += 1
    ws7.cell(row=r, column=1, value="Data Sources").font = SUBHEADER_FONT
    r += 1
    sources = [
        f"Cap Rates: {cap_data.get('source', 'N/A')}",
        f"Demographics: {demo.get('source', 'N/A')}",
        f"Rent Trends: {rent_data.get('source', 'N/A')}",
    ]
    for src in sources:
        ws7.cell(row=r, column=1, value=src).font = NORMAL_FONT
        r += 1

    # NOI line chart
    if pf:
        r += 1
        ws7.cell(row=r, column=1, value="Year")
        ws7.cell(row=r, column=2, value="NOI")
        for i, yr in enumerate(pf):
            ws7.cell(row=r+1+i, column=1, value=f"Yr {yr['year']}")
            ws7.cell(row=r+1+i, column=2, value=yr["noi"])
        chart = LineChart()
        chart.title = "NOI Growth Trajectory"
        chart.style = 10
        chart.width = 18
        chart.height = 10
        chart.add_data(Reference(ws7, min_col=2, min_row=r, max_row=r+len(pf)), titles_from_data=True)
        chart.set_categories(Reference(ws7, min_col=1, min_row=r+1, max_row=r+len(pf)))
        ws7.add_chart(chart, f"D{r}")
    _widths(ws7, 4)

    # ========== TAB 8: Assumptions ==========
    ws8 = wb.create_sheet("Assumptions")
    ws8.sheet_properties.tabColor = DARK_BLUE
    _title(ws8, 1, 1, "Underwriting Assumptions")

    r = 3
    sections = [
        ("Deal Inputs", [
            ("Property Type", deal["property_type"], None),
            ("Address", deal["address"], None),
            ("Year Built", deal["year_built"], None),
            ("Units", deal["total_units"], None),
            ("Total SF", deal["total_sf"], '#,##0'),
            ("In-Place Rent", deal["in_place_rent"], CURRENCY_FMT),
            ("Market Rent", deal["market_rent"], CURRENCY_FMT),
            ("Occupancy", deal["occupancy"], PCT_FMT),
            ("Current NOI", deal["current_noi"], CURRENCY_FMT),
        ]),
        ("Acquisition", [
            ("Purchase Price", deal["purchase_price"], CURRENCY_FMT),
            ("Closing Costs", deal["closing_costs_pct"], PCT_FMT),
            ("Deferred Maintenance", deal["deferred_maintenance"], CURRENCY_FMT),
            ("Planned CapEx", deal["planned_capex"], CURRENCY_FMT),
            ("Total Project Cost", derived["total_project_cost"], CURRENCY_FMT),
        ]),
        ("Financing", [
            ("LTV", deal["ltv"], PCT_FMT),
            ("Interest Rate", deal["interest_rate"], PCT_FMT),
            ("Amortization", f"{deal['amortization_years']} years", None),
            ("Loan Term", f"{deal['loan_term_years']} years", None),
            ("IO Period", f"{deal['io_period_years']} years", None),
            ("Loan Amount", derived["loan_amount"], CURRENCY_FMT),
            ("Equity Required", derived["equity_required"], CURRENCY_FMT),
        ]),
        ("Operating", [
            ("Revenue Growth", deal["revenue_growth_rate"], PCT_FMT),
            ("Expense Growth", deal["expense_growth_rate"], PCT_FMT),
            ("Management Fee", deal["management_fee_pct"], PCT_FMT),
            ("Reserves/Unit", deal["replacement_reserves_per_unit"], CURRENCY_FMT),
            ("Expense Ratio", derived["expense_ratio"], PCT_FMT),
        ]),
        ("Exit", [
            ("Hold Period", f"{deal['hold_period_years']} years", None),
            ("Exit Cap Spread", deal["exit_cap_rate_spread"], PCT_FMT),
            ("Sale Costs", deal["sale_costs_pct"], PCT_FMT),
            ("Going-In Cap Rate", derived["going_in_cap_rate"], PCT_FMT),
            ("Exit Cap Rate", derived["exit_cap_rate"], PCT_FMT),
        ]),
    ]

    for section_name, items in sections:
        ws8.cell(row=r, column=1, value=section_name).font = HEADER_FONT
        ws8.cell(row=r, column=1).fill = HEADER_FILL
        ws8.cell(row=r, column=2).fill = HEADER_FILL
        r += 1
        for label, val, fmt in items:
            ws8.cell(row=r, column=1, value=label).font = NORMAL_FONT
            c = ws8.cell(row=r, column=2, value=val)
            c.font = NORMAL_FONT
            if fmt:
                c.number_format = fmt
            r += 1
        r += 1
    _widths(ws8, 2, 20)

    # ========== TAB 9: ML Valuation (conditional) ==========
    if ml_valuation and not ml_valuation.get("error"):
        ws9 = wb.create_sheet("ML Valuation")
        ws9.sheet_properties.tabColor = GOLD
        _title(ws9, 1, 1, "ML-Based Property Valuation")

        r = 3
        ws9.cell(row=r, column=1, value="Assessment").font = SUBHEADER_FONT
        ws9.cell(row=r, column=2, value=ml_valuation["assessment"]).font = Font(name="Calibri", bold=True, size=14, color=GOLD)
        r += 2

        val_items = [
            ("Predicted Value/Unit", ml_valuation["predicted_value_per_unit"], CURRENCY_FMT),
            ("Predicted Total Value", ml_valuation["predicted_total_value"], CURRENCY_FMT),
            ("Actual Price/Unit", ml_valuation["actual_price_per_unit"], CURRENCY_FMT),
            ("Premium/Discount", ml_valuation["premium_discount_pct"] / 100, PCT_FMT),
            ("Train R² Score", ml_valuation.get("train_r2", ml_valuation.get("r2_score")), '0.0000'),
            ("Test R² Score (held-out)", ml_valuation.get("test_r2"), '0.0000'),
            ("Test MAE ($)", ml_valuation.get("test_mae"), CURRENCY_FMT),
            ("Test MAPE (%)", ml_valuation.get("test_mape"), '0.0"%"'),
            ("Model Type", ml_valuation["model_type"], None),
            ("Training Samples", ml_valuation["training_samples"], '#,##0'),
            ("Test Samples", ml_valuation.get("test_samples", 100), '#,##0'),
            ("Features Used", ml_valuation["features_used"], None),
            ("Data Note", "Trained on SYNTHETIC data calibrated to FRED/Census", None),
        ]
        for label, val, fmt in val_items:
            ws9.cell(row=r, column=1, value=label).font = NORMAL_FONT
            c = ws9.cell(row=r, column=2, value=val)
            c.font = NORMAL_FONT
            if fmt:
                c.number_format = fmt
            r += 1

        # Feature importances
        r += 1
        ws9.cell(row=r, column=1, value="Feature Importances").font = SUBHEADER_FONT
        r += 1
        for c_idx, h in enumerate(["Feature", "Importance"], 1):
            ws9.cell(row=r, column=c_idx, value=h)
        _hdr_row(ws9, r, 2)
        r += 1
        for feat, imp in ml_valuation["feature_importances"].items():
            ws9.cell(row=r, column=1, value=feat.replace("_", " ").title()).font = NORMAL_FONT
            ws9.cell(row=r, column=2, value=imp).number_format = '0.0000'
            _data_row(ws9, r, 2)
            r += 1

        _widths(ws9, 2, 22)

    # ========== TAB 10: Lease Analysis (conditional) ==========
    if lease_analysis and not lease_analysis.get("error"):
        ws10 = wb.create_sheet("Lease Analysis")
        ws10.sheet_properties.tabColor = GOLD
        _title(ws10, 1, 1, "Lease Document Analysis")

        r = 3
        if lease_analysis.get("summary"):
            c = ws10.cell(row=r, column=1, value=lease_analysis["summary"])
            c.font = NORMAL_FONT
            c.alignment = Alignment(wrap_text=True)
            ws10.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            r += 2

        ws10.cell(row=r, column=1, value="Key Lease Terms").font = SUBHEADER_FONT
        r += 1
        for c_idx, h in enumerate(["Term", "Value"], 1):
            ws10.cell(row=r, column=c_idx, value=h)
        _hdr_row(ws10, r, 2)
        r += 1

        term_fields = [
            ("Tenant", "tenant_name"),
            ("Landlord", "landlord_name"),
            ("Lease Type", "lease_type"),
            ("Monthly Rent", "monthly_rent"),
            ("Annual Rent", "annual_rent"),
            ("Rent/SF", "rent_per_sf"),
            ("Term (months)", "lease_term_months"),
            ("Start Date", "lease_start_date"),
            ("End Date", "lease_end_date"),
            ("Escalation", "escalation_clause"),
            ("Annual Escalation %", "annual_escalation_pct"),
            ("Renewal Options", "renewal_options"),
            ("Security Deposit", "security_deposit"),
            ("TI Allowance", "ti_allowance"),
            ("CAM Charges", "cam_charges"),
            ("Permitted Use", "permitted_use"),
        ]
        for label, key in term_fields:
            val = lease_analysis.get(key)
            if val is not None:
                ws10.cell(row=r, column=1, value=label).font = NORMAL_FONT
                c = ws10.cell(row=r, column=2, value=val)
                c.font = NORMAL_FONT
                if isinstance(val, (int, float)) and val > 100:
                    c.number_format = CURRENCY_FMT
                _data_row(ws10, r, 2)
                r += 1

        # Key clauses
        clauses = lease_analysis.get("key_clauses", [])
        if clauses:
            r += 1
            ws10.cell(row=r, column=1, value="Key Clauses").font = SUBHEADER_FONT
            r += 1
            for clause in clauses:
                ws10.cell(row=r, column=1, value=clause).font = NORMAL_FONT
                r += 1

        # Risk flags
        flags = lease_analysis.get("risk_flags", [])
        if flags:
            r += 1
            ws10.cell(row=r, column=1, value="Risk Flags").font = Font(name="Calibri", bold=True, size=11, color="CC0000")
            r += 1
            for flag in flags:
                ws10.cell(row=r, column=1, value=flag).font = NORMAL_FONT
                ws10.cell(row=r, column=1).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                r += 1

        ws10.cell(row=r+1, column=1, value=f"Analysis method: {lease_analysis.get('analysis_method', 'N/A')}").font = Font(name="Calibri", size=9, italic=True)
        _widths(ws10, 3, 25)

    # ========== TAB 11: Rent Forecast (conditional) ==========
    if rent_prediction and not rent_prediction.get("error"):
        ws11 = wb.create_sheet("Rent Forecast")
        ws11.sheet_properties.tabColor = GOLD
        _title(ws11, 1, 1, "Predictive Rent Growth Model")

        r = 3
        meta = [
            ("Method", rent_prediction["method"]),
            ("Data Source", rent_prediction["data_source"]),
            ("Training Points", rent_prediction["training_points"]),
            ("Hold Period", f"{rent_prediction['hold_period']} years"),
            ("Current Rent/Unit", rent_prediction["current_rent"]),
            ("Avg Predicted Growth", f"{rent_prediction['avg_predicted_growth']}%"),
            ("Historical Avg Growth", f"{rent_prediction['historical_avg']}%"),
        ]
        for label, val in meta:
            ws11.cell(row=r, column=1, value=label).font = SUBHEADER_FONT
            c = ws11.cell(row=r, column=2, value=val)
            c.font = NORMAL_FONT
            if isinstance(val, (int, float)) and val > 100:
                c.number_format = CURRENCY_FMT
            r += 1

        # Forecast table
        r += 1
        hold = rent_prediction["hold_period"]
        hdrs = [""] + [f"Year {i+1}" for i in range(hold)]
        for c_idx, h in enumerate(hdrs, 1):
            ws11.cell(row=r, column=c_idx, value=h)
        _hdr_row(ws11, r, len(hdrs))
        r += 1

        ws11.cell(row=r, column=1, value="Growth Rate").font = NORMAL_FONT
        for i, rate in enumerate(rent_prediction["predicted_rates"]):
            ws11.cell(row=r, column=i+2, value=rate/100).number_format = PCT_FMT
        _data_row(ws11, r, len(hdrs))
        r += 1

        ws11.cell(row=r, column=1, value="Rent/Unit").font = NORMAL_FONT
        for i, rent in enumerate(rent_prediction["predicted_rents_per_unit"]):
            ws11.cell(row=r, column=i+2, value=rent).number_format = CURRENCY_FMT
        _data_row(ws11, r, len(hdrs), alt=True)
        r += 1

        ws11.cell(row=r, column=1, value="Annual Revenue").font = SUBHEADER_FONT
        for i, rev_val in enumerate(rent_prediction["predicted_annual_revenue"]):
            ws11.cell(row=r, column=i+2, value=rev_val).number_format = CURRENCY_FMT
        r += 1

        # Historical rates
        r += 1
        ws11.cell(row=r, column=1, value="Historical Growth Rates (Recent)").font = SUBHEADER_FONT
        r += 1
        for i, rate in enumerate(rent_prediction["historical_rates"]):
            ws11.cell(row=r, column=1, value=f"Period {i+1}").font = NORMAL_FONT
            ws11.cell(row=r, column=2, value=rate/100).number_format = PCT_FMT
            r += 1

        # Rent forecast chart
        r += 1
        chart_start = r
        ws11.cell(row=r, column=1, value="Year")
        ws11.cell(row=r, column=2, value="Predicted Rent/Unit")
        for i, rent in enumerate(rent_prediction["predicted_rents_per_unit"]):
            ws11.cell(row=r+1+i, column=1, value=f"Year {i+1}")
            ws11.cell(row=r+1+i, column=2, value=rent)

        chart = LineChart()
        chart.title = "Predicted Rent Growth"
        chart.style = 10
        chart.width = 18
        chart.height = 10
        chart.add_data(Reference(ws11, min_col=2, min_row=chart_start, max_row=chart_start+hold), titles_from_data=True)
        chart.set_categories(Reference(ws11, min_col=1, min_row=chart_start+1, max_row=chart_start+hold))
        ws11.add_chart(chart, f"D3")

        _widths(ws11, len(hdrs), 14)
        ws11.column_dimensions["A"].width = 22

    # Save
    filename = f"underwriting_{job_id}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)
    return filepath
